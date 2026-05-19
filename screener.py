"""
Bulk stock screener: pulls fundamentals for S&P 500 + Russell 1000,
stores everything in SQLite, exports a color-coded Excel dashboard
highlighting "growth with a bit of risk" candidates.
"""

import json
import os
import sqlite3
import subprocess
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from io import StringIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import pandas as pd
import yfinance as yf

# Make yfinance impersonate a real Chrome browser to bypass "Invalid Crumb" 401s
try:
    from curl_cffi import requests as _curl_requests
    _YF_SESSION = _curl_requests.Session(impersonate="chrome")
except Exception:
    _YF_SESSION = None
    print("(curl_cffi not installed — yfinance may hit 401 Invalid Crumb errors. "
          "Run: pip install curl_cffi)")

WIKI_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0 Safari/537.36"
)


def _read_wiki_tables(url: str) -> list[pd.DataFrame]:
    """pd.read_html(url) hits HTTP 403 on Wikipedia; fetch with a UA first."""
    req = Request(url, headers={"User-Agent": WIKI_UA})
    with urlopen(req, timeout=30) as resp:
        html = resp.read().decode("utf-8", errors="ignore")
    return pd.read_html(StringIO(html))
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
TICKERS_FILE = ROOT / "tickers.txt"
DB_FILE = ROOT / "screener_data.db"
EXCEL_FILE = ROOT / f"screener_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

MAX_WORKERS = 20          # concurrent yfinance fetches — 20 is the sweet spot
PROGRESS_EVERY = 100      # print a progress line every N tickers
CACHE_HOURS = 12          # reuse last fetch if newer than this (set 0 to disable)
CACHE_FILE = Path(__file__).parent / "_fetch_cache_v5_edgar.parquet"

# ---------- SEC EDGAR config ----------
USE_EDGAR = True                                          # set False to skip XBRL layer
EDGAR_CACHE_DIR = Path(__file__).parent / "_edgar_cache"  # per-company JSON dump
EDGAR_CACHE_DAYS = 30                                     # filings rarely change
# SEC requires a real User-Agent identifying you. EDIT THIS to your name + email:
EDGAR_UA = os.environ.get("EDGAR_UA", "Stock Screener brody@example.com")

# "Growth with a bit of risk" thresholds — set any to None to DISABLE that filter.
# Defaults are loose; tighten any of these as you learn which stocks survive.
FILTERS = {
    # ---- Core growth / valuation ----
    "revenue_growth_min": 0.15,      # 15% YoY (was 20%)
    "forward_pe_max": 50.0,          # was 40
    "beta_min": 1.0,                 # was 1.3
    "beta_max": 2.5,                 # was 2.2
    "debt_to_equity_max": 200.0,     # was 150
    "gross_margin_min": 0.20,        # was 0.25
    "market_cap_min": 500_000_000,   # was 1B
    "profit_margin_min": None,       # set 0.05 for >=5% net margin
    "earnings_growth_min": None,     # set 0.10 for >=10% EPS growth
    "return_on_equity_min": None,    # set 0.12 for >=12% ROE
    # ---- "Don't chase the rally" momentum guards ----
    "max_pct_above_52w_low": 1.50,   # was 1.0 (now allows up to 150% off low)
    "min_pct_below_52w_high": 0.05,  # was 0.08
    "short_float_max": None,         # set 0.10 to avoid crowded shorts
    "insider_ownership_min": None,   # set 0.02 for >=2% insider ownership
    # ---- Statement-derived quality gates (all OPT-IN now) ----
    "fcf_conversion_min": None,      # set 0.80 for FCF >=80% of net income
    "fcf_margin_min": None,          # set 0.05 for ≥5% FCF margin
    "accrual_ratio_max": 0.15,       # was 0.10 — looser earnings-quality gate
    "roic_min": None,                # set 0.08 for ≥8% ROIC
    "roic_wacc_spread_min": None,    # set 0.02 for ROIC>WACC by 2%
    "gross_margin_trend_min": None,  # set 0.00 to require improving GM
    "op_margin_min": None,           # set 0.08 for >=8% operating margin
    "op_margin_trend_min": None,     # set 0.00 to require improving OM
    "interest_coverage_min": None,   # set 3.0 to require 3x coverage
    "current_ratio_min": None,       # set 1.0
    "net_debt_to_fcf_max": None,     # set 4.0 to avoid levered stories
    "goodwill_pct_assets_max": None, # set 0.50
    "sbc_pct_revenue_max": None,     # set 0.20
    "rule_of_40_min": None,          # set 0.40 for SaaS-style screen
    "rev_cagr_3y_min": None,         # set 0.10 to require durable growth
    "ev_to_fcf_max": None,           # set 35.0 for FCF valuation discipline
    "ev_to_ebitda_max": None,        # set 25.0 for EBITDA valuation discipline
    "capex_to_revenue_max": None,    # set 0.15 to avoid capex-heavy names
    "sec_backlog_to_rev_min": None,  # set 0.50 for backlog-supported demand
}


# ---------- Ticker universe ----------

def fetch_ticker_universe() -> list[str]:
    """Pull S&P 500 from Wikipedia and merge with a Russell-1000 proxy."""
    print("Fetching S&P 500 from Wikipedia...")
    sp500 = _read_wiki_tables(
        "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"
    )[0]["Symbol"].str.replace(".", "-", regex=False).tolist()

    print("Fetching Russell 1000 from Wikipedia...")
    try:
        r1000 = _read_wiki_tables(
            "https://en.wikipedia.org/wiki/Russell_1000_Index"
        )
        # The constituents table varies in position; find the one with a Ticker/Symbol column
        r1000_syms = []
        for tbl in r1000:
            cols = [c.lower() for c in tbl.columns.astype(str)]
            for c in ("symbol", "ticker"):
                if c in cols:
                    col = tbl.columns[cols.index(c)]
                    r1000_syms = tbl[col].astype(str).str.replace(".", "-", regex=False).tolist()
                    break
            if r1000_syms:
                break
    except Exception as e:
        print(f"  Russell 1000 fetch failed ({e}); using S&P 500 only.")
        r1000_syms = []

    universe = sorted(set(sp500) | set(r1000_syms))
    # Strip obvious junk
    universe = [t for t in universe if t and t.isascii() and len(t) <= 6]
    return universe


def load_or_build_tickers() -> list[str]:
    if TICKERS_FILE.exists() and TICKERS_FILE.stat().st_size > 0:
        tickers = [
            line.strip().upper()
            for line in TICKERS_FILE.read_text().splitlines()
            if line.strip() and not line.startswith("#")
        ]
        print(f"Loaded {len(tickers)} tickers from {TICKERS_FILE.name}")
        return tickers

    print("No tickers.txt found — building universe...")
    tickers = fetch_ticker_universe()
    TICKERS_FILE.write_text("\n".join(tickers))
    print(f"Wrote {len(tickers)} tickers to {TICKERS_FILE.name}")
    return tickers


# ---------- Data fetch ----------

FIELDS = [
    "symbol", "shortName", "sector", "industry", "currentPrice",
    "marketCap", "enterpriseValue", "sharesOutstanding",
    "forwardPE", "trailingPE", "beta",
    "revenueGrowth", "earningsGrowth", "grossMargins", "profitMargins",
    "debtToEquity", "returnOnEquity", "freeCashflow",
    "fiftyTwoWeekHigh", "fiftyTwoWeekLow",
    "shortPercentOfFloat", "heldPercentInsiders",
    "longBusinessSummary",
]


# ---------- Theme dictionary ----------
# Map common natural-language themes to keyword sets.
# Anything the user types that isn't in here is used as a literal substring match.
THEMES = {
    "data center": [
        "data center", "data centre", "hyperscale", "colocation",
        "server", "cloud infrastructure", "edge computing",
    ],
    "data center construction": [
        "data center", "data centre", "hyperscale", "engineering",
        "construction", "electrical contractor", "modular", "infrastructure",
    ],
    "thermal management": [
        "thermal", "cooling", "liquid cooling", "hvac", "heat exchange",
        "chiller", "precision cooling",
    ],
    "solar": [
        "solar", "photovoltaic", "pv ", "renewable energy",
    ],
    "solar materials": [
        "polysilicon", "wafer", "photovoltaic", "solar cell",
        "solar module", "encapsulant", "ingot",
    ],
    "solar panel direct materials": [
        "polysilicon", "wafer", "ingot", "solar cell", "solar module",
        "encapsulant", "glass", "silver paste", "photovoltaic",
    ],
    "semiconductor": [
        "semiconductor", "chip", "wafer", "fabless", "foundry", "lithography",
    ],
    "ai": [
        "artificial intelligence", "machine learning", "generative ai",
        "gpu", "accelerator", "inference", "large language model",
    ],
    "energy": [
        "energy", "power generation", "utility", "grid", "electricity",
        "natural gas", "renewable",
    ],
    "battery": [
        "battery", "lithium", "energy storage", "cathode", "anode",
    ],
    "uranium": [
        "uranium", "nuclear", "enrichment",
    ],
    "construction": [
        "construction", "engineering", "infrastructure", "contractor",
        "civil", "building",
    ],
    "defense": [
        "defense", "defence", "military", "aerospace", "missile", "weapons",
    ],
    "biotech": [
        "biotech", "therapeutics", "clinical", "pharmaceutical", "oncology",
    ],
    "cybersecurity": [
        "cybersecurity", "security software", "endpoint", "firewall",
        "threat detection", "zero trust",
    ],
    "quantum": [
        "quantum computing", "quantum",
    ],
    "robotics": [
        "robotics", "automation", "industrial robot", "autonomous",
    ],
    "ev": [
        "electric vehicle", "ev ", "charging", "powertrain",
    ],
    "fintech": [
        "fintech", "payments", "digital banking", "neobank",
    ],
}


def expand_themes(user_input: str) -> list[str]:
    """Turn 'data center construction, solar materials' into a flat keyword list."""
    keywords: list[str] = []
    for raw in user_input.split(","):
        theme = raw.strip().lower()
        if not theme:
            continue
        if theme in THEMES:
            keywords.extend(THEMES[theme])
        else:
            # Unknown theme — use the literal phrase as a keyword
            keywords.append(theme)
    # De-dupe while preserving order
    seen = set()
    out = []
    for k in keywords:
        if k not in seen:
            seen.add(k)
            out.append(k)
    return out


def apply_theme_filter(df: pd.DataFrame, keywords: list[str]) -> pd.DataFrame:
    """Keep rows where any keyword appears in sector/industry/business summary."""
    if not keywords:
        return df
    haystack = (
        df["sector"].fillna("").astype(str).str.lower()
        + " || " + df["industry"].fillna("").astype(str).str.lower()
        + " || " + df["longBusinessSummary"].fillna("").astype(str).str.lower()
        + " || " + df["shortName"].fillna("").astype(str).str.lower()
    )
    pattern = "|".join(pd.Series(keywords).str.lower().map(lambda s: s.strip()))
    return df[haystack.str.contains(pattern, regex=True, na=False)]


def prompt_for_themes() -> list[str]:
    """Interactive prompt at the start of each run."""
    # Skip prompt for scheduled / non-interactive runs
    if os.environ.get("SCREENER_NO_OPEN") == "1" or not sys.stdin.isatty():
        env_themes = os.environ.get("SCREENER_THEMES", "").strip()
        if env_themes:
            print(f"Using themes from SCREENER_THEMES env: {env_themes}")
            return expand_themes(env_themes)
        return []
    print()
    print("=" * 70)
    print("THEME FILTER (optional)")
    print("=" * 70)
    print("Describe the kinds of stocks you want, comma-separated.")
    print("Examples:")
    print("  data center construction, thermal management, solar materials")
    print("  ai, semiconductor, energy")
    print("  uranium, defense")
    print()
    print(f"Known themes: {', '.join(sorted(THEMES.keys()))}")
    print("(Unknown phrases are matched as literal text. Press ENTER to skip.)")
    print()
    raw = input("Themes> ").strip()
    if not raw:
        print("No theme filter — showing all sectors.\n")
        return []
    kws = expand_themes(raw)
    print(f"Matching against {len(kws)} keywords: {kws}\n")
    return kws


def safe_get(info: dict, key: str):
    v = info.get(key)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    return v


# ---------- Hedge-fund-grade KPIs from real financial statements ----------
# yfinance has renamed these rows across versions, so we try multiple aliases.
ROW_ALIASES = {
    "ocf": [
        "Operating Cash Flow", "Total Cash From Operating Activities",
        "Cash Flow From Continuing Operating Activities",
        "Cash Flow From Operating Activities",
    ],
    "capex": ["Capital Expenditure", "Capital Expenditures"],
    "revenue": ["Total Revenue", "Revenue", "Operating Revenue"],
    "cogs": ["Cost Of Revenue", "Reconciled Cost Of Revenue", "Cost Of Goods Sold"],
    "gross_profit": ["Gross Profit"],
    "rnd": ["Research And Development", "Research Development"],
    "net_income": [
        "Net Income", "Net Income Common Stockholders",
        "Net Income From Continuing Operations",
        "Net Income From Continuing Operation Net Minority Interest",
    ],
    "total_assets": ["Total Assets"],
    "current_assets": ["Current Assets", "Total Current Assets"],
    "current_liab": ["Current Liabilities", "Total Current Liabilities"],
    "cash": [
        "Cash And Cash Equivalents",
        "Cash Cash Equivalents And Short Term Investments",
        "Cash",
    ],
    "operating_income": ["Operating Income", "EBIT"],
    "ebitda": ["EBITDA", "Normalized EBITDA"],
    "total_debt": ["Total Debt", "Net Debt"],
    "equity": [
        "Stockholders Equity", "Total Stockholders Equity",
        "Common Stock Equity",
    ],
    "goodwill": ["Goodwill"],
    "intangibles": ["Other Intangible Assets", "Goodwill And Other Intangible Assets"],
    "interest_expense": ["Interest Expense", "Interest Expense Non Operating"],
    "receivables": ["Accounts Receivable", "Receivables", "Net Receivables"],
    "inventory": ["Inventory", "Inventories"],
    "payables": ["Accounts Payable", "Payables"],
    "depreciation": [
        "Reconciled Depreciation", "Depreciation And Amortization",
        "Depreciation Amortization Depletion",
    ],
    "sbc": ["Stock Based Compensation"],
    "buybacks": ["Repurchase Of Capital Stock", "Common Stock Repurchase"],
    "deferred_rev": ["Current Deferred Revenue", "Deferred Revenue"],
}

# Assumptions for WACC estimation
RISK_FREE_RATE = 0.045   # ~10Y Treasury
EQUITY_RISK_PREMIUM = 0.055
TAX_RATE = 0.21


def _row(df, key: str):
    """Return a Series for a logical row (most-recent first) or None."""
    if df is None or getattr(df, "empty", True):
        return None
    for label in ROW_ALIASES.get(key, []):
        if label in df.index:
            return df.loc[label]
    return None


def _safe_div(a, b):
    try:
        if b is None or a is None or b == 0 or pd.isna(b) or pd.isna(a):
            return None
        return float(a) / float(b)
    except Exception:
        return None


def _series_val(s, i=0, default=None):
    """Safely fetch s.iloc[i] as float or default."""
    try:
        if s is None or len(s) <= i:
            return default
        v = s.iloc[i]
        if v is None or pd.isna(v):
            return default
        return float(v)
    except Exception:
        return default


def deep_kpis(t, info: dict) -> dict:
    """Compute statement-derived KPIs that .info can never give you."""
    is_ = getattr(t, "income_stmt", None)
    bs_ = getattr(t, "balance_sheet", None)
    cf_ = getattr(t, "cashflow", None)

    ocf   = _row(cf_, "ocf")
    cpx   = _row(cf_, "capex")            # yfinance reports negative
    rev   = _row(is_, "revenue")
    cogs  = _row(is_, "cogs")
    gp    = _row(is_, "gross_profit")
    rnd   = _row(is_, "rnd")
    ni    = _row(is_, "net_income")
    ta    = _row(bs_, "total_assets")
    ca    = _row(bs_, "current_assets")
    cl    = _row(bs_, "current_liab")
    cash  = _row(bs_, "cash")
    opi   = _row(is_, "operating_income")
    ebitda= _row(is_, "ebitda")
    debt  = _row(bs_, "total_debt")
    eq    = _row(bs_, "equity")
    gw    = _row(bs_, "goodwill")
    intg  = _row(bs_, "intangibles")
    intexp= _row(is_, "interest_expense")
    ar    = _row(bs_, "receivables")
    inv   = _row(bs_, "inventory")
    ap    = _row(bs_, "payables")
    dep   = _row(cf_, "depreciation")
    sbc   = _row(cf_, "sbc")
    buy   = _row(cf_, "buybacks")
    defrv = _row(bs_, "deferred_rev")

    out: dict = {}

    # ============ FREE CASH FLOW & MARGINS ============
    fcf0 = None
    if ocf is not None and cpx is not None:
        ocf0, cpx0 = _series_val(ocf), _series_val(cpx)
        if ocf0 is not None and cpx0 is not None:
            fcf0 = ocf0 + cpx0  # capex is negative in yfinance
            out["fcf_latest"] = fcf0
            out["fcf_margin"] = _safe_div(fcf0, _series_val(rev))
            out["fcf_conversion"] = _safe_div(fcf0, _series_val(ni))
            out["owner_earnings"] = ocf0 + cpx0  # OCF - maintenance CapEx proxy

    # ============ MARGIN TRENDS (5Y slope, positive = improving) ============
    def _trend(series):
        """Slope of % change per year, normalized to base."""
        if series is None or len(series) < 3:
            return None
        try:
            vals = [float(v) for v in series.dropna()]
            if len(vals) < 3 or vals[-1] == 0:
                return None
            return (vals[0] - vals[-1]) / abs(vals[-1]) / (len(vals) - 1)
        except Exception:
            return None

    if rev is not None and cogs is not None:
        gm_series = (rev - cogs) / rev
        out["gross_margin_trend"] = _trend(gm_series)
        out["gross_margin_latest"] = _series_val(gm_series)
    elif gp is not None and rev is not None:
        gm_series = gp / rev
        out["gross_margin_trend"] = _trend(gm_series)
        out["gross_margin_latest"] = _series_val(gm_series)

    if opi is not None and rev is not None:
        om_series = opi / rev
        out["op_margin_trend"] = _trend(om_series)
        out["op_margin_latest"] = _series_val(om_series)

    # ============ ACCRUAL RATIO (earnings quality) ============
    out["accrual_ratio"] = _safe_div(
        (_series_val(ni) or 0) - (_series_val(ocf) or 0), _series_val(ta)
    ) if ni is not None and ocf is not None else None

    # ============ ROIC / TANGIBLE / INCREMENTAL ============
    if opi is not None and ta is not None and cl is not None:
        nopat0 = (_series_val(opi) or 0) * (1 - TAX_RATE)
        cash0 = _series_val(cash) or 0
        ic0 = (_series_val(ta) or 0) - (_series_val(cl) or 0) - cash0
        out["roic"] = _safe_div(nopat0, ic0)

        # Incremental ROIC (YoY)
        nopat1 = (_series_val(opi, 1) or 0) * (1 - TAX_RATE)
        ic1 = ((_series_val(ta, 1) or 0) - (_series_val(cl, 1) or 0)
               - (_series_val(cash, 1) or 0))
        d_nopat, d_ic = nopat0 - nopat1, ic0 - ic1
        if abs(d_ic) > 1:
            out["incremental_roic"] = d_nopat / d_ic

        # Tangible ROIC
        intang0 = (_series_val(gw) or 0) + (_series_val(intg) or 0)
        out["tangible_roic"] = _safe_div(nopat0, ic0 - intang0)

    # ============ WACC ESTIMATE & SPREAD ============
    beta = info.get("beta")
    mcap = info.get("marketCap")
    debt0 = _series_val(debt) or 0
    if beta and mcap and (mcap + debt0) > 0:
        cost_equity = RISK_FREE_RATE + float(beta) * EQUITY_RISK_PREMIUM
        cost_debt = _safe_div(abs(_series_val(intexp) or 0), debt0) or 0.05
        e_weight = mcap / (mcap + debt0)
        d_weight = debt0 / (mcap + debt0)
        wacc = e_weight * cost_equity + d_weight * cost_debt * (1 - TAX_RATE)
        out["wacc"] = wacc
        if "roic" in out and out["roic"] is not None:
            out["roic_wacc_spread"] = out["roic"] - wacc

    # ============ EFFICIENCY & CAPITAL ALLOCATION ============
    out["asset_turnover"] = _safe_div(_series_val(rev), _series_val(ta))
    out["sbc_pct_revenue"] = _safe_div(_series_val(sbc), _series_val(rev))
    out["rnd_pct_revenue"] = _safe_div(_series_val(rnd), _series_val(rev))
    out["reinvestment_rate"] = _safe_div(abs(_series_val(cpx) or 0), _series_val(ocf))
    if buy is not None and mcap:
        out["buyback_yield"] = _safe_div(abs(_series_val(buy) or 0), mcap)

    # ============ CASH CONVERSION CYCLE ============
    cogs0 = _series_val(cogs)
    rev0 = _series_val(rev)
    if rev0 and cogs0:
        dso = _safe_div(_series_val(ar), rev0)
        dio = _safe_div(_series_val(inv), cogs0)
        dpo = _safe_div(_series_val(ap), cogs0)
        if dso is not None and dio is not None and dpo is not None:
            out["ccc_days"] = (dso + dio - dpo) * 365

    # ============ DEFERRED REVENUE GROWTH ============
    if defrv is not None and len(defrv) >= 2:
        d0, d1 = _series_val(defrv, 0), _series_val(defrv, 1)
        out["deferred_rev_growth"] = _safe_div((d0 or 0) - (d1 or 0), d1)

    # ============ INCREMENTAL OPERATING MARGIN ============
    if opi is not None and rev is not None and len(opi) > 1 and len(rev) > 1:
        d_rev = (_series_val(rev) or 0) - (_series_val(rev, 1) or 0)
        d_op  = (_series_val(opi) or 0) - (_series_val(opi, 1) or 0)
        if abs(d_rev) > 1:
            out["incremental_margin"] = d_op / d_rev

    # ============ MAINTENANCE vs GROWTH CAPEX ============
    dep0 = _series_val(dep)
    cpx_abs = abs(_series_val(cpx) or 0)
    if dep0 is not None and cpx_abs:
        out["maintenance_capex"] = min(dep0, cpx_abs)
        out["growth_capex"] = max(0.0, cpx_abs - dep0)
        out["growth_capex_pct"] = _safe_div(out["growth_capex"], cpx_abs)

    # ============ BALANCE SHEET STRENGTH ============
    out["current_ratio"] = _safe_div(_series_val(ca), _series_val(cl))
    out["goodwill_pct_assets"] = _safe_div(_series_val(gw), _series_val(ta))
    if fcf0 and debt0:
        out["net_debt_to_fcf"] = _safe_div(debt0 - (_series_val(cash) or 0), fcf0)
    out["interest_coverage"] = _safe_div(_series_val(opi), abs(_series_val(intexp) or 0))

    # ============ REVENUE CAGR (3Y, 5Y) ============
    # CAGR is only mathematically valid when both endpoint values are POSITIVE.
    # A fractional power of a negative number = complex result → crashes storage.
    def _cagr(end, start, years):
        try:
            if end is None or start is None or years <= 0:
                return None
            if end <= 0 or start <= 0:
                return None
            return (end / start) ** (1.0 / years) - 1
        except Exception:
            return None

    if rev is not None and len(rev) >= 2:
        for years_back, label in [(2, "rev_cagr_3y"), (4, "rev_cagr_5y")]:
            if len(rev) > years_back:
                out[label] = _cagr(_series_val(rev), _series_val(rev, years_back), years_back)
    if rev is not None and len(rev) >= 2:
        years = len(rev) - 1
        out["rev_cagr_stmt"] = _cagr(_series_val(rev), _series_val(rev, -1), years)

    # ============ MARKET EXPECTATIONS ============
    ev = info.get("enterpriseValue")
    if ev and fcf0:
        out["ev_to_fcf"] = ev / fcf0
    if ebitda is not None and ev:
        out["ev_to_ebitda"] = _safe_div(ev, _series_val(ebitda))
    # Rule of 40: revenue growth + FCF margin
    rg = info.get("revenueGrowth")
    if rg is not None and "fcf_margin" in out and out["fcf_margin"] is not None:
        out["rule_of_40"] = float(rg) + out["fcf_margin"]
    # PEG on FCF: (Price / FCF per share) / growth%
    shares = info.get("sharesOutstanding")
    price = info.get("currentPrice")
    if shares and price and fcf0 and rg and rg > 0:
        fcf_per_share = fcf0 / shares
        if fcf_per_share > 0:
            p_fcf = price / fcf_per_share
            out["peg_on_fcf"] = p_fcf / (rg * 100)

    # ============ BOOK-TO-BILL (proxy via deferred rev / revenue) ============
    # True book-to-bill needs order data; deferred revenue growth is a soft proxy.
    if defrv is not None and rev0:
        out["book_to_bill_proxy"] = _safe_div(_series_val(defrv), rev0)

    # ============ CAPEX-TO-REVENUE ============
    out["capex_to_revenue"] = _safe_div(cpx_abs, rev0)

    return out


# ============================================================
# SEC EDGAR XBRL — institutional data, free, straight from filings
# ============================================================

_edgar_lock = threading.Lock()
_edgar_last_call = [0.0]   # mutable container so threads share it
_EDGAR_MIN_INTERVAL = 0.11 # ~9 req/sec, safely under SEC's 10/sec cap
_ticker_to_cik: dict[str, str] | None = None


def _edgar_throttled_get(url: str) -> bytes:
    """Thread-safe rate-limited HTTPS GET with SEC-required UA."""
    with _edgar_lock:
        wait = _EDGAR_MIN_INTERVAL - (time.time() - _edgar_last_call[0])
        if wait > 0:
            time.sleep(wait)
        _edgar_last_call[0] = time.time()
    req = Request(url, headers={"User-Agent": EDGAR_UA, "Accept-Encoding": "gzip"})
    with urlopen(req, timeout=30) as resp:
        data = resp.read()
        if resp.headers.get("Content-Encoding") == "gzip":
            import gzip
            data = gzip.decompress(data)
        return data


def _load_ticker_cik_map() -> dict[str, str]:
    """Map TICKER → 10-digit zero-padded CIK. Cached for 30 days."""
    global _ticker_to_cik
    if _ticker_to_cik is not None:
        return _ticker_to_cik

    cache = EDGAR_CACHE_DIR / "ticker_cik_map.json"
    EDGAR_CACHE_DIR.mkdir(exist_ok=True)
    fresh = (cache.exists() and
             datetime.now() - datetime.fromtimestamp(cache.stat().st_mtime)
             < timedelta(days=EDGAR_CACHE_DAYS))
    if not fresh:
        try:
            raw = _edgar_throttled_get("https://www.sec.gov/files/company_tickers.json")
            cache.write_bytes(raw)
        except Exception as e:
            print(f"  EDGAR ticker map fetch failed: {e}")
            if not cache.exists():
                _ticker_to_cik = {}
                return _ticker_to_cik

    raw_map = json.loads(cache.read_text())
    mapping: dict[str, str] = {}
    for v in raw_map.values():
        mapping[v["ticker"].upper()] = str(v["cik_str"]).zfill(10)
    _ticker_to_cik = mapping
    return mapping


def _fetch_company_facts(cik: str) -> dict | None:
    """Pull /companyfacts/CIK*.json with disk cache."""
    EDGAR_CACHE_DIR.mkdir(exist_ok=True)
    cache = EDGAR_CACHE_DIR / f"CIK{cik}.json"
    fresh = (cache.exists() and
             datetime.now() - datetime.fromtimestamp(cache.stat().st_mtime)
             < timedelta(days=EDGAR_CACHE_DAYS))
    if fresh:
        try:
            return json.loads(cache.read_text())
        except Exception:
            pass
    try:
        raw = _edgar_throttled_get(
            f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
        )
        cache.write_bytes(raw)
        return json.loads(raw)
    except HTTPError as e:
        if e.code == 404:
            cache.write_text("{}")  # negative-cache so we don't retry
            return None
        raise


def _latest_annual(facts: dict, tag: str, unit: str = "USD"):
    """Most recent annual (FY) value for a us-gaap or dei tag. Returns (value, end_date)."""
    for ns in ("us-gaap", "ifrs-full", "dei"):
        node = facts.get("facts", {}).get(ns, {}).get(tag)
        if not node:
            continue
        units = node.get("units", {})
        # USD usually; some tags are pure (no unit) or shares
        for u in (unit, "USD/shares", "shares", "pure"):
            entries = units.get(u, [])
            annual = [e for e in entries if e.get("fp") == "FY" and e.get("form", "").startswith("10-K")]
            if not annual:
                annual = [e for e in entries if e.get("fp") == "FY"]
            if not annual:
                continue
            annual.sort(key=lambda e: e.get("end", ""), reverse=True)
            return annual[0].get("val"), annual[0].get("end")
    return None, None


def _sum_last_year_quarterly(facts: dict, tag: str) -> float | None:
    """Sum the most recent 4 quarterly values for tags only reported quarterly."""
    for ns in ("us-gaap", "ifrs-full"):
        node = facts.get("facts", {}).get(ns, {}).get(tag)
        if not node:
            continue
        entries = node.get("units", {}).get("USD", [])
        q = [e for e in entries if e.get("fp") in ("Q1", "Q2", "Q3", "Q4")]
        if len(q) < 4:
            continue
        q.sort(key=lambda e: e.get("end", ""), reverse=True)
        try:
            return sum(float(e["val"]) for e in q[:4])
        except Exception:
            continue
    return None


# Tag groups — try alternatives in order until one returns data
EDGAR_TAGS = {
    "revenue":            ["Revenues", "RevenueFromContractWithCustomerExcludingAssessedTax",
                           "SalesRevenueNet", "RevenueFromContractWithCustomerIncludingAssessedTax"],
    "sbc":                ["ShareBasedCompensation",
                           "AllocatedShareBasedCompensationExpense"],
    "rnd":                ["ResearchAndDevelopmentExpense"],
    "capex":              ["PaymentsToAcquirePropertyPlantAndEquipment",
                           "PaymentsToAcquireProductiveAssets"],
    "depreciation":       ["DepreciationDepletionAndAmortization",
                           "DepreciationAndAmortization", "Depreciation"],
    "goodwill":           ["Goodwill"],
    "intangibles":        ["IntangibleAssetsNetExcludingGoodwill",
                           "FiniteLivedIntangibleAssetsNet"],
    "total_assets":       ["Assets"],
    "current_liab":       ["LiabilitiesCurrent"],
    "lease_liab":         ["OperatingLeaseLiability",
                           "OperatingLeaseLiabilityCurrent"],
    "long_term_debt":     ["LongTermDebt", "LongTermDebtNoncurrent"],
    "debt_due_1y":        ["LongTermDebtMaturitiesRepaymentsOfPrincipalInNextTwelveMonths",
                           "LongTermDebtCurrent"],
    "debt_due_2y":        ["LongTermDebtMaturitiesRepaymentsOfPrincipalInYearTwo"],
    "debt_due_5y":        ["LongTermDebtMaturitiesRepaymentsOfPrincipalInYearFive"],
    "deferred_rev":       ["ContractWithCustomerLiability",
                           "DeferredRevenue",
                           "ContractWithCustomerLiabilityCurrent"],
    "backlog":            ["RevenueRemainingPerformanceObligation"],
    "acquisition_spend":  ["BusinessCombinationConsiderationTransferred",
                           "PaymentsToAcquireBusinessesNetOfCashAcquired"],
}


def _get_tag(facts: dict, key: str) -> tuple[float | None, str | None]:
    for tag in EDGAR_TAGS.get(key, []):
        v, end = _latest_annual(facts, tag)
        if v is not None:
            return float(v), end
    return None, None


def edgar_kpis(symbol: str) -> dict:
    """Pull SEC XBRL facts and compute filings-derived KPIs."""
    out: dict = {}
    cik = _load_ticker_cik_map().get(symbol.upper())
    if not cik:
        return {"sec_status": "no_cik"}  # not a US registrant
    facts = _fetch_company_facts(cik)
    if not facts or "facts" not in facts:
        return {"sec_status": "no_facts"}

    rev,   _ = _get_tag(facts, "revenue")
    sbc,   _ = _get_tag(facts, "sbc")
    rnd,   _ = _get_tag(facts, "rnd")
    cpx,   _ = _get_tag(facts, "capex")
    dep,   _ = _get_tag(facts, "depreciation")
    gw,    _ = _get_tag(facts, "goodwill")
    intg,  _ = _get_tag(facts, "intangibles")
    ta,    _ = _get_tag(facts, "total_assets")
    cl,    _ = _get_tag(facts, "current_liab")
    lease, _ = _get_tag(facts, "lease_liab")
    ltd,   _ = _get_tag(facts, "long_term_debt")
    due1,  _ = _get_tag(facts, "debt_due_1y")
    due2,  _ = _get_tag(facts, "debt_due_2y")
    due5,  _ = _get_tag(facts, "debt_due_5y")
    defrv, _ = _get_tag(facts, "deferred_rev")
    backlog,end = _get_tag(facts, "backlog")
    acq,   _ = _get_tag(facts, "acquisition_spend")

    out["sec_status"]            = "ok"
    out["sec_revenue"]           = rev
    out["sec_sbc"]               = sbc
    out["sec_sbc_pct_revenue"]   = _safe_div(sbc, rev)
    out["sec_rnd_pct_revenue"]   = _safe_div(rnd, rev)
    out["sec_capex_intensity"]   = _safe_div(cpx, rev)
    # Maintenance vs growth CapEx (using D&A as maintenance proxy)
    if dep is not None and cpx is not None:
        out["sec_maint_capex"]   = min(dep, cpx)
        out["sec_growth_capex"]  = max(0.0, cpx - dep)
        out["sec_growth_capex_pct"] = _safe_div(out["sec_growth_capex"], cpx)
    # Tangible capital base
    if ta is not None:
        intang = (gw or 0) + (intg or 0)
        out["sec_tangible_assets"] = ta - intang
        out["sec_goodwill_pct"]    = _safe_div(gw, ta)
    # Lease-adjusted debt
    if ltd is not None or lease is not None:
        out["sec_total_debt_w_leases"] = (ltd or 0) + (lease or 0)
    # Debt-maturity wall (% due within 1 year)
    total_known_debt = (due1 or 0) + (due2 or 0) + (due5 or 0)
    if total_known_debt > 0 and due1 is not None:
        out["sec_pct_debt_due_1y"] = due1 / ((ltd or total_known_debt) or 1)
    # Backlog / revenue (the "real" book-to-bill for industrials & SaaS)
    if backlog is not None:
        out["sec_backlog"]       = backlog
        out["sec_backlog_to_rev"] = _safe_div(backlog, rev)
        out["sec_backlog_date"]  = end
    # Deferred revenue growth — uses XBRL history if available
    out["sec_deferred_rev"]      = defrv
    # M&A intensity — acquisition spend vs revenue
    out["sec_acq_intensity"]     = _safe_div(acq, rev)
    # Organic growth proxy: if you grew revenue without significant acquisitions, it's organic
    if rev and acq is not None:
        out["sec_acq_pct_rev"]   = _safe_div(acq, rev)

    return out


def fetch_one(sym: str) -> dict:
    """Fetch a single ticker's fundamentals + statement KPIs. Never raises."""
    try:
        t = yf.Ticker(sym, session=_YF_SESSION) if _YF_SESSION else yf.Ticker(sym)
        info = t.info or {}
        if not info.get("marketCap") and not info.get("shortName"):
            return {"symbol": sym, "error": "no_data"}
        row = {f: safe_get(info, f) for f in FIELDS} | {"symbol": sym}
        try:
            row.update(deep_kpis(t, info))
        except Exception as e:
            row["_deep_error"] = f"{type(e).__name__}: {str(e)[:80]}"
        if USE_EDGAR:
            try:
                row.update(edgar_kpis(sym))
            except Exception as e:
                row["sec_status"] = f"err:{type(e).__name__}"
        return row
    except Exception as e:
        return {"symbol": sym, "error": f"{type(e).__name__}: {str(e)[:100]}"}


def _load_cache() -> pd.DataFrame | None:
    if CACHE_HOURS <= 0 or not CACHE_FILE.exists():
        return None
    age = datetime.now() - datetime.fromtimestamp(CACHE_FILE.stat().st_mtime)
    if age > timedelta(hours=CACHE_HOURS):
        return None
    try:
        df = pd.read_parquet(CACHE_FILE)
        print(f"Using cached fetch from {age.total_seconds() / 3600:.1f}h ago "
              f"({len(df)} rows). Delete {CACHE_FILE.name} to force refresh.")
        return df
    except Exception:
        return None


def fetch_all(tickers: list[str]) -> pd.DataFrame:
    cached = _load_cache()
    if cached is not None:
        return cached

    total = len(tickers)
    all_rows: list[dict] = []
    start = time.time()
    print(f"Fetching {total} tickers with {MAX_WORKERS} parallel workers...")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(fetch_one, sym): sym for sym in tickers}
        for i, fut in enumerate(as_completed(futures), start=1):
            all_rows.append(fut.result())
            if i % PROGRESS_EVERY == 0 or i == total:
                rate = i / (time.time() - start)
                eta = (total - i) / rate if rate else 0
                print(f"  {i}/{total}  ({rate:.1f}/s, ~{eta:.0f}s left)")

    df = pd.DataFrame(all_rows)
    df["fetched_at"] = datetime.now().isoformat(timespec="seconds")
    elapsed = time.time() - start
    print(f"Fetch done in {elapsed:.1f}s ({total / elapsed:.1f} tickers/s).")

    try:
        sanitize_for_storage(df).to_parquet(CACHE_FILE, index=False)
        print(f"Cached fetch to {CACHE_FILE.name}")
    except Exception as e:
        print(f"  (cache write skipped: {e})")
    return df


# ---------- Persistence ----------

_BAD_STRINGS = {"infinity", "-infinity", "inf", "-inf", "nan", "none", "<na>", "nat", ""}


def _coerce(v):
    """Coerce one cell to a SQLite/parquet-safe scalar."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return None if v.strip().lower() in _BAD_STRINGS else v
    if isinstance(v, float):
        if pd.isna(v) or v == float("inf") or v == float("-inf"):
            return None
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, complex):
        # Complex numbers can't be stored; take real part if imaginary is ~0, else None
        return float(v.real) if abs(v.imag) < 1e-9 else None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return str(v)


def sanitize_for_storage(df: pd.DataFrame) -> pd.DataFrame:
    """Make a DataFrame safe for both parquet and SQLite — drops inf, fixes types."""
    safe = df.copy()

    for col in safe.columns:
        dtype = safe[col].dtype
        dtype_str = str(dtype)
        # Stringify what storage can't handle directly
        if any(k in dtype_str for k in
               ("datetime", "timedelta", "category", "interval", "period")):
            safe[col] = (safe[col].astype(str)
                         .replace({"NaT": None, "nan": None, "<NA>": None, "None": None}))
            continue
        # Pandas nullable extension types — convert to plain object then resolve below
        if dtype_str in ("Int64", "Float64", "boolean", "string"):
            safe[col] = safe[col].astype(object)
            dtype = safe[col].dtype

        if safe[col].dtype == object:
            # Sanitize cell contents (Timestamps/lists/Infinity/etc.)
            safe[col] = safe[col].map(_coerce)
            non_null = safe[col].dropna()
            if len(non_null) > 0:
                numeric = pd.to_numeric(safe[col], errors="coerce")
                # Replace any +/-inf that slipped through with NaN
                numeric = numeric.replace([float("inf"), float("-inf")], pd.NA)
                if numeric.notna().sum() >= len(non_null) * 0.8:
                    safe[col] = numeric.astype("float64")
                else:
                    safe[col] = safe[col].map(lambda v: None if v is None else str(v))
        elif "float" in dtype_str:
            safe[col] = safe[col].replace([float("inf"), float("-inf")], pd.NA)
        elif "complex" in dtype_str:
            # Convert complex column to real part, drop columns with non-trivial imag
            safe[col] = safe[col].apply(
                lambda v: None if v is None or pd.isna(v)
                else (float(v.real) if abs(getattr(v, "imag", 0)) < 1e-9 else None)
            )

    return safe


def save_to_sqlite(df: pd.DataFrame) -> None:
    safe = sanitize_for_storage(df)
    with sqlite3.connect(DB_FILE) as conn:
        # Detect schema mismatch with prior runs and rebuild if columns changed
        existing_cols = set()
        cur = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='fundamentals'"
        )
        if cur.fetchone():
            existing_cols = {
                row[1] for row in conn.execute("PRAGMA table_info(fundamentals)")
            }
            if existing_cols != set(safe.columns):
                print("  Schema changed since last run — archiving old table.")
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                conn.execute(f"ALTER TABLE fundamentals RENAME TO fundamentals_{stamp}")

        safe.to_sql("fundamentals", conn, if_exists="append", index=False)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_symbol_time "
            "ON fundamentals(symbol, fetched_at)"
        )
    print(f"Appended {len(safe)} rows to {DB_FILE.name}")


# ---------- Screening + Excel ----------

def apply_filters(
    df: pd.DataFrame,
    verbose: bool = True,
    filters: dict | None = None,
) -> pd.DataFrame:
    f = FILTERS.copy()
    if filters:
        f.update(filters)
    df = df.copy()
    funnel: list[tuple[str, int]] = [("Starting universe", len(df))]

    def _step(mask: pd.Series, label: str) -> pd.Series:
        nonlocal df
        survivors = int(mask.sum())
        funnel.append((label, survivors))
        return mask

    # Derived momentum columns — how far from 52w extremes
    df["pct_above_52w_low"] = (
        (df["currentPrice"] - df["fiftyTwoWeekLow"]) / df["fiftyTwoWeekLow"]
    )
    df["pct_below_52w_high"] = (
        (df["fiftyTwoWeekHigh"] - df["currentPrice"]) / df["fiftyTwoWeekHigh"]
    )

    mask = pd.Series(True, index=df.index)

    def _gate(condition: pd.Series, label: str):
        nonlocal mask
        mask = mask & condition
        funnel.append((label, int(mask.sum())))

    _gate(pd.to_numeric(df["revenueGrowth"], errors="coerce") >= f["revenue_growth_min"],
          f"Rev growth ≥ {f['revenue_growth_min']:.0%}")
    _gate((pd.to_numeric(df["forwardPE"], errors="coerce") > 0) &
          (pd.to_numeric(df["forwardPE"], errors="coerce") <= f["forward_pe_max"]),
          f"Fwd P/E ≤ {f['forward_pe_max']}")
    _gate(pd.to_numeric(df["beta"], errors="coerce").between(f["beta_min"], f["beta_max"]),
          f"Beta {f['beta_min']}–{f['beta_max']}")
    _gate(pd.to_numeric(df["debtToEquity"], errors="coerce") <= f["debt_to_equity_max"],
          f"D/E ≤ {f['debt_to_equity_max']}")
    _gate(pd.to_numeric(df["grossMargins"], errors="coerce") >= f["gross_margin_min"],
          f"Gross margin ≥ {f['gross_margin_min']:.0%}")
    _gate(pd.to_numeric(df["marketCap"], errors="coerce") >= f["market_cap_min"],
          f"Market cap ≥ ${f['market_cap_min']/1e9:.1f}B")
    _gate(pd.to_numeric(df["pct_above_52w_low"], errors="coerce") <= f["max_pct_above_52w_low"],
          f"≤ {f['max_pct_above_52w_low']:.0%} above 52w low")
    _gate(pd.to_numeric(df["pct_below_52w_high"], errors="coerce") >= f["min_pct_below_52w_high"],
          f"≥ {f['min_pct_below_52w_high']:.0%} below 52w high")

    # Optional quality gates (skipped if threshold is None). NaN values pass here:
    # this keeps otherwise strong names from being thrown out just because one
    # vendor metric is missing.
    for col, key, op in [
        ("profitMargins",      "profit_margin_min",       "ge"),
        ("earningsGrowth",     "earnings_growth_min",     "ge"),
        ("returnOnEquity",     "return_on_equity_min",    "ge"),
        ("shortPercentOfFloat","short_float_max",         "le"),
        ("heldPercentInsiders","insider_ownership_min",   "ge"),
        ("fcf_margin",         "fcf_margin_min",          "ge"),
        ("fcf_conversion",     "fcf_conversion_min",      "ge"),
        ("accrual_ratio",      "accrual_ratio_max",       "le"),
        ("roic",               "roic_min",                "ge"),
        ("roic_wacc_spread",   "roic_wacc_spread_min",    "ge"),
        ("gross_margin_trend", "gross_margin_trend_min",  "ge"),
        ("op_margin_latest",   "op_margin_min",           "ge"),
        ("op_margin_trend",    "op_margin_trend_min",     "ge"),
        ("interest_coverage",  "interest_coverage_min",   "ge"),
        ("current_ratio",      "current_ratio_min",       "ge"),
        ("net_debt_to_fcf",    "net_debt_to_fcf_max",     "le"),
        ("goodwill_pct_assets","goodwill_pct_assets_max", "le"),
        ("sbc_pct_revenue",    "sbc_pct_revenue_max",     "le"),
        ("rule_of_40",         "rule_of_40_min",          "ge"),
        ("rev_cagr_3y",        "rev_cagr_3y_min",         "ge"),
        ("ev_to_fcf",          "ev_to_fcf_max",           "le"),
        ("ev_to_ebitda",       "ev_to_ebitda_max",        "le"),
        ("capex_to_revenue",   "capex_to_revenue_max",    "le"),
        ("sec_backlog_to_rev", "sec_backlog_to_rev_min",  "ge"),
    ]:
        thresh = f.get(key)
        if thresh is None or col not in df.columns:
            continue
        s = pd.to_numeric(df[col], errors="coerce")
        cond = (s >= thresh) if op == "ge" else (s <= thresh)
        _gate(cond | s.isna(), f"{col} {op} {thresh}")

    if verbose:
        print("\n=== FILTER FUNNEL ===")
        for label, n in funnel:
            print(f"  {n:5d}  {label}")

    # Composite quality + growth + momentum + valuation score
    out = df[mask].copy()
    def _num(col, default=0):
        return pd.to_numeric(out.get(col, default), errors="coerce").fillna(default)

    growth     = _num("revenueGrowth")
    off_high   = _num("pct_below_52w_high")
    fcf_m      = _num("fcf_margin")
    roic_v     = _num("roic")
    spread     = _num("roic_wacc_spread")
    gm_trend   = _num("gross_margin_trend")
    rule40     = _num("rule_of_40")
    ev_fcf     = _num("ev_to_fcf", 50).clip(lower=1)  # cheaper EV/FCF = better
    accruals   = _num("accrual_ratio").abs()

    # Higher is better for each multiplier; clip to keep extremes from dominating
    out["value_score"] = (
        growth.clip(-0.5, 2.0)
        * off_high.clip(0, 0.5)
        * (1 + fcf_m.clip(-0.2, 0.5))
        * (1 + roic_v.clip(-0.2, 0.5))
        * (1 + spread.clip(-0.1, 0.3))
        * (1 + gm_trend.clip(-0.2, 0.2))
        * (1 + rule40.clip(-0.5, 1.0))
        / (1 + ev_fcf / 25)
        / (1 + accruals * 5)
    )
    return out.sort_values("value_score", ascending=False)


# Column display config: (header_label, number_format, width, conditional_rule)
# Rules: "growth_color" green-high, "pe_color" low-good, "momentum_color" mid-good,
#        "data_bar" simple bar, None = no rule.
COLUMN_CONFIG = {
    "symbol":              ("Ticker",       "@",             10, None),
    "shortName":           ("Company",      "@",             28, None),
    "sector":              ("Sector",       "@",             18, None),
    "industry":            ("Industry",     "@",             28, None),
    "currentPrice":        ("Price",        "$#,##0.00",     11, None),
    "marketCap":           ("Market Cap",   '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "forwardPE":           ("Fwd P/E",      "0.00",          10, "pe_color"),
    "trailingPE":          ("TTM P/E",      "0.00",          10, "pe_color"),
    "beta":                ("Beta",         "0.00",           8, "momentum_color"),
    "revenueGrowth":       ("Rev Growth",   "0.0%",          12, "growth_color"),
    "earningsGrowth":      ("EPS Growth",   "0.0%",          12, "growth_color"),
    "grossMargins":        ("Gross Margin", "0.0%",          12, "growth_color"),
    "profitMargins":       ("Profit Margin","0.0%",          13, "growth_color"),
    "debtToEquity":        ("D/E",          "0.0",            8, "pe_color"),
    "returnOnEquity":      ("ROE",          "0.0%",           9, "growth_color"),
    "freeCashflow":        ("Free Cashflow",'_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "pct_below_52w_high":  ("% Off 52w Hi", "0.0%",          13, "growth_color"),
    "pct_above_52w_low":   ("% Up Fr 52wLo","0.0%",          13, "pe_color"),
    "fiftyTwoWeekHigh":    ("52w High",     "$#,##0.00",     11, None),
    "fiftyTwoWeekLow":     ("52w Low",      "$#,##0.00",     11, None),
    "value_score":         ("Score",        "0.000",          9, "growth_color"),
    "enterpriseValue":     ("EV",           '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "sharesOutstanding":   ("Shares Out",   "#,##0",         14, None),
    "shortPercentOfFloat": ("Short % Float","0.0%",          12, "pe_color"),
    "heldPercentInsiders": ("Insider Own",  "0.0%",          12, "growth_color"),
    # --- Statement-derived KPIs: Cash Flow ---
    "fcf_latest":          ("FCF (latest)", '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "fcf_margin":          ("FCF Margin",   "0.0%",          11, "growth_color"),
    "fcf_conversion":      ("FCF/NI",       "0.0%",          10, "growth_color"),
    "owner_earnings":      ("Owner Earn",   '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "accrual_ratio":       ("Accruals",     "0.00%",         10, "pe_color"),
    # --- Profitability Trends ---
    "gross_margin_latest": ("Gross Margin", "0.0%",          12, "growth_color"),
    "gross_margin_trend":  ("GM Trend",     "0.0%",          10, "growth_color"),
    "op_margin_latest":    ("Op Margin",    "0.0%",          11, "growth_color"),
    "op_margin_trend":     ("OM Trend",     "0.0%",          10, "growth_color"),
    # --- Returns ---
    "roic":                ("ROIC",         "0.0%",           9, "growth_color"),
    "wacc":                ("WACC",         "0.0%",           9, "pe_color"),
    "roic_wacc_spread":    ("ROIC-WACC",    "0.0%",          11, "growth_color"),
    "incremental_roic":    ("Δ ROIC",       "0.0%",          10, "growth_color"),
    "tangible_roic":       ("Tang ROIC",    "0.0%",          11, "growth_color"),
    # --- Efficiency ---
    "asset_turnover":      ("Asset Turn",   "0.00",          11, "growth_color"),
    "sbc_pct_revenue":     ("SBC %",        "0.0%",          10, "pe_color"),
    "rnd_pct_revenue":     ("R&D %",        "0.0%",          10, "growth_color"),
    "reinvestment_rate":   ("Reinvest %",   "0.0%",          11, "growth_color"),
    "buyback_yield":       ("Buyback Y",    "0.0%",          11, "growth_color"),
    "ccc_days":            ("CCC Days",     "0",             10, "pe_color"),
    "deferred_rev_growth": ("Def Rev Gr",   "0.0%",          11, "growth_color"),
    "incremental_margin":  ("Δ Op Margin",  "0.0%",          12, "growth_color"),
    "maintenance_capex":   ("Maint CapEx",  '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, None),
    "growth_capex":        ("Growth CapEx", '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, None),
    "growth_capex_pct":    ("Growth CpX%",  "0.0%",          12, "growth_color"),
    "capex_to_revenue":    ("CapEx/Rev",    "0.0%",          11, "pe_color"),
    # --- Balance Sheet ---
    "current_ratio":       ("Current Rt",   "0.00",          11, "growth_color"),
    "goodwill_pct_assets": ("GW % Assets",  "0.0%",          12, "pe_color"),
    "net_debt_to_fcf":     ("NetDebt/FCF",  "0.0",           12, "pe_color"),
    "interest_coverage":   ("Int Coverage", "0.0",           12, "growth_color"),
    # --- Growth ---
    "rev_cagr_3y":         ("Rev CAGR 3Y",  "0.0%",          12, "growth_color"),
    "rev_cagr_5y":         ("Rev CAGR 5Y",  "0.0%",          12, "growth_color"),
    "rev_cagr_stmt":       ("Rev CAGR All", "0.0%",          12, "growth_color"),
    # --- Market Expectations ---
    "ev_to_fcf":           ("EV/FCF",       "0.0",           10, "pe_color"),
    "ev_to_ebitda":        ("EV/EBITDA",    "0.0",           11, "pe_color"),
    "rule_of_40":          ("Rule of 40",   "0.0%",          11, "growth_color"),
    "peg_on_fcf":          ("PEG (FCF)",    "0.00",          10, "pe_color"),
    "book_to_bill_proxy":  ("Book/Bill*",   "0.00",          11, "growth_color"),
    # --- SEC EDGAR XBRL (institutional data) ---
    "sec_status":          ("SEC",          "@",              8, None),
    "sec_revenue":         ("SEC Rev",      '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "sec_sbc":             ("SEC SBC",      '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, None),
    "sec_sbc_pct_revenue": ("SEC SBC %",    "0.0%",          11, "pe_color"),
    "sec_rnd_pct_revenue": ("SEC R&D %",    "0.0%",          11, "growth_color"),
    "sec_capex_intensity": ("SEC CpX/Rev",  "0.0%",          12, "pe_color"),
    "sec_maint_capex":     ("Maint CpX",    '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, None),
    "sec_growth_capex":    ("Growth CpX",   '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, None),
    "sec_growth_capex_pct":("Gr CpX %",     "0.0%",          11, "growth_color"),
    "sec_tangible_assets": ("Tang Assets",  '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "sec_goodwill_pct":    ("GW % (SEC)",   "0.0%",          12, "pe_color"),
    "sec_total_debt_w_leases": ("Debt+Lease",'_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "pe_color"),
    "sec_pct_debt_due_1y": ("Debt Due 1Y",  "0.0%",          12, "pe_color"),
    "sec_backlog":         ("Backlog",      '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, "data_bar"),
    "sec_backlog_to_rev":  ("Backlog/Rev",  "0.00",          12, "growth_color"),
    "sec_backlog_date":    ("Backlog Date", "@",             13, None),
    "sec_deferred_rev":    ("Deferred Rev", '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)', 16, None),
    "sec_acq_intensity":   ("Acq/Rev",      "0.0%",          11, "pe_color"),
    "sec_acq_pct_rev":     ("Acq % Rev",    "0.0%",          11, "pe_color"),
    "fetched_at":          ("Fetched",      "@",             20, None),
    "longBusinessSummary": ("Business",     "@",             60, None),
    "_deep_error":         ("Stmt Error",   "@",             20, None),
    "error":               ("Error",        "@",             20, None),
}

# Preferred column order for the Hits sheet (most decision-relevant first)
HIT_COL_ORDER = [
    "symbol", "shortName", "sector", "industry", "currentPrice", "marketCap",
    "value_score",
    # ---------- Growth ----------
    "revenueGrowth", "rev_cagr_3y", "rev_cagr_5y", "earningsGrowth",
    "deferred_rev_growth",
    # ---------- Returns on Capital ----------
    "roic", "wacc", "roic_wacc_spread", "incremental_roic", "tangible_roic",
    # ---------- Cash Flow Quality ----------
    "fcf_margin", "fcf_conversion", "owner_earnings", "accrual_ratio",
    # ---------- Margin Trends ----------
    "gross_margin_latest", "gross_margin_trend",
    "op_margin_latest", "op_margin_trend", "incremental_margin",
    # ---------- Efficiency / Capital Allocation ----------
    "asset_turnover", "reinvestment_rate", "buyback_yield",
    "sbc_pct_revenue", "rnd_pct_revenue",
    "capex_to_revenue", "maintenance_capex", "growth_capex", "growth_capex_pct",
    "ccc_days",
    # ---------- Balance Sheet ----------
    "current_ratio", "net_debt_to_fcf", "interest_coverage",
    "goodwill_pct_assets", "debtToEquity",
    # ---------- Market Expectations / Valuation ----------
    "forwardPE", "ev_to_fcf", "ev_to_ebitda", "rule_of_40", "peg_on_fcf",
    "book_to_bill_proxy",
    # ---------- Sentiment / Ownership ----------
    "shortPercentOfFloat", "heldPercentInsiders", "beta",
    # ---------- Momentum ----------
    "pct_below_52w_high", "pct_above_52w_low",
    "fiftyTwoWeekHigh", "fiftyTwoWeekLow",
    # ---------- SEC EDGAR XBRL (institutional data) ----------
    "sec_status",
    "sec_backlog", "sec_backlog_to_rev", "sec_backlog_date",
    "sec_sbc_pct_revenue", "sec_rnd_pct_revenue", "sec_capex_intensity",
    "sec_growth_capex_pct", "sec_maint_capex", "sec_growth_capex",
    "sec_goodwill_pct", "sec_tangible_assets",
    "sec_total_debt_w_leases", "sec_pct_debt_due_1y",
    "sec_acq_intensity", "sec_deferred_rev",
    # ---------- Other ----------
    "enterpriseValue", "fcf_latest", "freeCashflow",
    "grossMargins", "profitMargins", "returnOnEquity",
    "longBusinessSummary",
]


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Apply header style, formats, conditional formatting, freeze, autofilter."""
    n_rows = len(df)
    n_cols = len(df.columns)
    if n_rows == 0 or n_cols == 0:
        return

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header styling
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cfg = COLUMN_CONFIG.get(col_name)
        if cfg:
            cell.value = cfg[0]  # friendly label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Per-column formatting + conditional formatting
    for col_idx, col_name in enumerate(df.columns, start=1):
        cfg = COLUMN_CONFIG.get(col_name)
        if not cfg:
            continue
        _, num_fmt, width, rule = cfg
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width

        # Apply number format to data cells
        for row_idx in range(2, n_rows + 2):
            ws.cell(row=row_idx, column=col_idx).number_format = num_fmt

        # Conditional formatting
        rng = f"{letter}2:{letter}{n_rows + 1}"
        if rule == "growth_color":
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ))
        elif rule == "pe_color":  # low is good — reverse scale
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            ))
        elif rule == "momentum_color":  # mid is best (beta 1.5-1.8)
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="63BE7B",
                end_type="max", end_color="F8696B",
            ))
        elif rule == "data_bar":
            ws.conditional_formatting.add(rng, DataBarRule(
                start_type="min", end_type="max", color="638EC6",
                showValue=True,
            ))

    # Freeze the header row and enable autofilter
    ws.freeze_panes = "C2"  # freeze header + ticker + company columns
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"

    # Row striping for readability
    stripe_fill = PatternFill("solid", fgColor="F2F2F2")
    for row_idx in range(2, n_rows + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = stripe_fill

    ws.row_dimensions[1].height = 28


def _build_summary(hits: pd.DataFrame, full: pd.DataFrame, themes: list[str]) -> pd.DataFrame:
    """Top-level summary sheet."""
    rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Universe size", f"{len(full):,} stocks"),
        ("Stocks passing filter", f"{len(hits):,}"),
        ("Theme filter", ", ".join(themes) if themes else "(none — all sectors)"),
        ("", ""),
        ("--- ACTIVE FILTERS ---", ""),
    ]
    for k, v in FILTERS.items():
        rows.append((k, v))
    rows.append(("", ""))
    rows.append(("--- SECTOR BREAKDOWN OF HITS ---", ""))
    if len(hits):
        for sector, count in hits["sector"].fillna("(unknown)").value_counts().items():
            rows.append((sector, count))
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _available_excel_path(path: Path) -> Path:
    """Return a non-existing Excel path near the requested path."""
    if not path.exists():
        return path
    for i in range(1, 100):
        candidate = path.with_name(f"{path.stem}_{i}.xlsx")
        if not candidate.exists():
            return candidate
    return path.with_name(f"{path.stem}_{datetime.now():%H%M%S}.xlsx")


def export_excel(full: pd.DataFrame, hits: pd.DataFrame, themes: list[str]) -> Path:
    # Reorder hits columns for analysis-friendly layout
    hit_cols = [c for c in HIT_COL_ORDER if c in hits.columns]
    extra = [c for c in hits.columns if c not in hit_cols]
    hits_view = hits[hit_cols + extra]

    summary = _build_summary(hits, full, themes)
    output_file = _available_excel_path(EXCEL_FILE)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        hits_view.to_excel(writer, sheet_name="Screened Hits", index=False)
        full.to_excel(writer, sheet_name="All Data", index=False)

        wb = writer.book

        # Style the Summary sheet
        sm = wb["Summary"]
        sm.column_dimensions["A"].width = 32
        sm.column_dimensions["B"].width = 40
        for cell in sm[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)

        _format_sheet(wb["Screened Hits"], hits_view)
        _format_sheet(wb["All Data"], full)

    print(f"Wrote {output_file.name} - {len(hits)} hits / {len(full)} total")
    return output_file


# ---------- Main ----------

def main() -> None:
    # Ask up-front so the user can walk away during the long fetch
    theme_keywords = prompt_for_themes()

    tickers = load_or_build_tickers()
    df = fetch_all(tickers)

    save_to_sqlite(df)

    clean = df[df.get("error").isna()] if "error" in df.columns else df
    themed = apply_theme_filter(clean, theme_keywords)
    if theme_keywords:
        print(f"Theme filter: {len(themed)} of {len(clean)} stocks matched.")
    hits = apply_filters(themed)

    print("\n=== TOP 20 HITS ===")
    cols = ["symbol", "shortName", "sector", "revenueGrowth", "forwardPE",
            "beta", "pct_below_52w_high", "pct_above_52w_low", "marketCap"]
    print(hits[cols].head(20).to_string(index=False))

    excel_file = export_excel(clean, hits, theme_keywords)

    # Also dump a plain-text winners list for quick scanning
    winners_file = ROOT / "winners_latest.txt"
    winners_file.write_text(
        f"Top picks as of {datetime.now():%Y-%m-%d %H:%M}\n"
        + "=" * 60 + "\n"
        + hits[cols].head(30).to_string(index=False)
    )
    print(f"Wrote {winners_file.name}")

    # Auto-open the Excel dashboard on Windows
    if sys.platform == "win32" and os.environ.get("SCREENER_NO_OPEN") != "1":
        try:
            os.startfile(excel_file)
        except Exception as e:
            print(f"(Could not auto-open Excel: {e})")


if __name__ == "__main__":
    main()
